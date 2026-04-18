"""Parse CSV/XLSX staff uploads into normalized ``uploaded_datasets`` rows.

Raises ``DatasourceParseError`` with a structured ``detail`` payload on any
validation failure so callers can surface actionable messages via HTTP 422.
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field
from typing import Any

from openpyxl import load_workbook

from .validation import coerce_numeric, coerce_year, derive_state_fips, validate_metric_name


@dataclass
class MappingConfig:
    """Contributor-declared column mapping for a datasource upload."""

    geo_column: str
    metric_value_column: str
    metric_name: str
    race_column: str | None = None
    year_column: str | None = None
    # Fallback year used when ``year_column`` is None.
    data_year: int | None = None

    def __post_init__(self) -> None:
        self.metric_name = validate_metric_name(self.metric_name)


class DatasourceParseError(Exception):
    """Raised when a datasource upload fails parse/validation.

    ``detail`` is a JSON-serializable payload suitable for ``HTTPException(422, detail=...)``.
    """

    def __init__(self, message: str, *, detail: dict[str, Any] | None = None) -> None:
        super().__init__(message)
        self.detail = detail or {"message": message}


def read_csv_bytes(content: bytes) -> tuple[list[str], list[dict[str, str]]]:
    """Parse CSV bytes into ``(header, rows)``.

    - Decodes as ``utf-8-sig`` so a UTF-8 BOM is stripped from the first cell.
    - Trims whitespace from header names (cell values are left raw so downstream
      coercion can handle whitespace consistently).
    - Raises ``DatasourceParseError`` if the file is empty or has no header row.
    """
    if not content:
        raise DatasourceParseError("file is empty")
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    try:
        raw_header = next(reader)
    except StopIteration:
        raise DatasourceParseError("file has no header row")

    header = [h.strip() for h in raw_header]
    rows = [dict(zip(header, r)) for r in reader if any(cell.strip() for cell in r)]
    return header, rows


def read_xlsx_bytes(content: bytes) -> tuple[list[str], list[dict[str, str]]]:
    """Parse XLSX bytes into ``(header, rows)`` using openpyxl.

    - Reads the first worksheet only.
    - ``header`` is taken from row 1, trimmed of whitespace.
    - Empty rows (all cells None / blank) are skipped.
    - Non-string cell values are stringified so downstream coercion sees the
      same shape as the CSV reader.
    """
    if not content:
        raise DatasourceParseError("file is empty")
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise DatasourceParseError(f"could not open xlsx: {exc}") from exc

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        raw_header = next(rows_iter)
    except StopIteration:
        raise DatasourceParseError("file has no header row")

    header = [("" if h is None else str(h)).strip() for h in raw_header]
    if not any(header):
        raise DatasourceParseError("header row is empty")

    rows: list[dict[str, str]] = []
    for raw_row in rows_iter:
        if raw_row is None or all(cell is None or str(cell).strip() == "" for cell in raw_row):
            continue
        row = {h: ("" if v is None else str(v)) for h, v in zip(header, raw_row)}
        rows.append(row)
    return header, rows


MIN_VALID_ROWS = 10
MAX_BAD_FIPS_RATIO = 0.10
MIN_NUMERIC_RATIO = 0.90
PREVIEW_LIMIT = 20
SUPPORTED_EXTS = {".csv", ".xlsx"}


@dataclass
class ParseResult:
    normalized_rows: list[dict[str, Any]]
    row_count: int
    dropped_counts: dict[str, int] = field(default_factory=dict)
    preview_rows: list[dict[str, Any]] = field(default_factory=list)


def _check_columns_exist(header: list[str], mapping: MappingConfig) -> None:
    required = [mapping.geo_column, mapping.metric_value_column]
    optional = [c for c in (mapping.race_column, mapping.year_column) if c]
    missing = [c for c in required + optional if c not in header]
    if missing:
        raise DatasourceParseError(
            "declared columns not found in file header",
            detail={"missing_columns": missing, "header": header},
        )


def _normalize_rows(
    rows: list[dict[str, str]], mapping: MappingConfig
) -> tuple[list[dict[str, Any]], dict[str, int]]:
    """Return (normalized_rows, dropped_counts).

    Drops rows that fail FIPS, numeric, or year parsing individually; the caller
    enforces overall drop-ratio rules.
    """
    dropped = {"bad_fips": 0, "non_numeric": 0, "bad_year": 0}
    out: list[dict[str, Any]] = []
    for row in rows:
        try:
            geo_raw = row.get(mapping.geo_column)
            state_fips = derive_state_fips(geo_raw)
            geo_fips = str(geo_raw).strip() if geo_raw is not None else ""
        except ValueError:
            dropped["bad_fips"] += 1
            continue

        value_raw = row.get(mapping.metric_value_column)
        try:
            value = coerce_numeric(value_raw)
        except ValueError:
            dropped["non_numeric"] += 1
            continue

        if mapping.year_column:
            try:
                year = coerce_year(row.get(mapping.year_column))
            except ValueError:
                dropped["bad_year"] += 1
                continue
        else:
            if mapping.data_year is None:
                raise DatasourceParseError(
                    "data_year is required when year_column is not set",
                )
            year = mapping.data_year

        race = None
        if mapping.race_column:
            raw_race = row.get(mapping.race_column)
            race = str(raw_race).strip() if raw_race is not None and str(raw_race).strip() else None

        out.append({
            "geo_fips": geo_fips,
            "state_fips": state_fips,
            "race": race,
            "year": year,
            "value": value,
        })
    return out, dropped


def parse_datasource_file(
    content: bytes, ext: str, mapping: MappingConfig
) -> ParseResult:
    """Parse + validate + normalize a staff datasource upload.

    Raises ``DatasourceParseError`` on any unrecoverable condition (missing
    columns, >10% bad FIPS, <90% numeric, <10 valid rows post-drop, etc.).
    """
    ext = ext.lower()
    if ext == ".csv":
        header, rows = read_csv_bytes(content)
    elif ext == ".xlsx":
        header, rows = read_xlsx_bytes(content)
    else:
        raise DatasourceParseError(
            f"unsupported file type {ext!r}",
            detail={"allowed": sorted(SUPPORTED_EXTS)},
        )

    _check_columns_exist(header, mapping)

    normalized, dropped = _normalize_rows(rows, mapping)
    total_seen = len(rows)

    if total_seen > 0 and (dropped["bad_fips"] / total_seen) > MAX_BAD_FIPS_RATIO:
        raise DatasourceParseError(
            "too many rows have unparseable FIPS values",
            detail={"dropped": {
                "reason": "bad_fips",
                "count": dropped["bad_fips"],
                "total": total_seen,
            }},
        )

    # Numeric rule is computed against rows that passed FIPS, to avoid double-counting.
    fips_ok = total_seen - dropped["bad_fips"]
    if fips_ok > 0:
        numeric_ratio = 1 - (dropped["non_numeric"] / fips_ok)
        if numeric_ratio < MIN_NUMERIC_RATIO:
            raise DatasourceParseError(
                "too many rows have non-numeric values in the metric column",
                detail={"dropped": {
                    "reason": "non_numeric",
                    "count": dropped["non_numeric"],
                    "total": fips_ok,
                }},
            )

    if len(normalized) < MIN_VALID_ROWS:
        raise DatasourceParseError(
            f"fewer than {MIN_VALID_ROWS} valid rows remained after validation",
            detail={"reason": "too_few_rows", "valid": len(normalized)},
        )

    return ParseResult(
        normalized_rows=normalized,
        row_count=len(normalized),
        dropped_counts=dropped,
        preview_rows=normalized[:PREVIEW_LIMIT],
    )
